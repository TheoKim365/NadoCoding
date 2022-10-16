from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells('b2:d2')
ws['b2'].value = 'Merged cell'

wb.save('sample_merge.xlsx')