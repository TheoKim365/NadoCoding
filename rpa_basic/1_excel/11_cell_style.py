from tkinter.tix import Tree
from turtle import color
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook('sample.xlsx')
ws = wb.active

# 번호, 영어, 수학
a1 = ws['A1'] # 번호
b1 = ws['B1'] # 영어
c1 = ws['C1'] # 수학 

# A열의 너비를 5로 설정 
ws.column_dimensions['A'].width = 5
# 1행의 높이를 5로 설정 
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color='ff0000', italic=True, bold=True)
b1.font = Font(color='cc33ff', name='Arial', strike=True)
c1.font = Font(color='0000ff', size=20, underline='single')

# 테두리 적용 
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 80점 넘는 셀에 대해서 초록색 적용 
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

        if cell.column == 1:  # A 번호열은 제외 
            continue
        
        # 셀이 정수형 데이터이고 80점 보다 높으면 
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = PatternFill(fgColor='00ff00', fill_type='solid')
            cell.font = Font(color='ff0000')

# 틀 고정
ws.freeze_panes = 'b2'

wb.save('sample_style.xlsx')