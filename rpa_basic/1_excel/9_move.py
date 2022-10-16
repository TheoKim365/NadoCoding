from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

# 번호, 영어, 수학
# 번호, (국어), 영어, 수학 
# ws.move_range('B1:C11', rows=0, cols=1)
# ws['B1'].value = '국어'

# 번호 영어 수학
# ws.move_range('C1:C11', rows=5, cols=-1)  # Formulae and references will not be updated.
# ws.insert_cols(2)  # Formulae and references will not be updated.

min_row = ws.min_row
max_row = ws.max_row
range = 'c' + str(min_row) + ':' + 'c' + str(max_row)
ws.move_range(range, cols=1)


wb.save('sample_korean.xlsx')