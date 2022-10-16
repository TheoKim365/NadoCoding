from openpyxl import load_workbook

wb = load_workbook('sample_merge.xlsx')
ws = wb.active

# b2:d2 병합되어 있던 셀을 해제
ws.unmerge_cells('b2:d2')

wb.save('sample_unmerge.xlsx')