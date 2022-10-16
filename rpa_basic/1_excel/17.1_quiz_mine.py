from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 현재까지 작성된 최종 성적을 데이터에 넣기
ws.append(['학번','출석','퀴즈1','퀴즈2','중간고사','기말공사','프로젝트'])
max_col = ws.max_column

scores = [
(1,	10,	8,	5,	15,	26,	12),
(2,	7,	3,	7,	15,	24,	18),
(3,	9,	5,	8,	8,	12,	4),
(4,	7,	8,	7,	17,	21,	18),
(5,	7,	8,	7,	16,	25,	15),
(6,	3,	5,	8,	8,	17,	0),
(7,	4,	9,	10,	16,	27,	18),
(8,	6,	6,	6,	15,	19,	17),
(9,	10,	10,	9,	19,	30,	19),
(10,	9,	8,	8,	20,	25,	20)
]

for s in scores:
    ws.append(s)

# 1. 퀴즈2 점수를 10으로 수정 
for idx, cell in enumerate(ws['d']):
    if idx == 0:  # 제목인 경우 skip
        continue
    cell.value = 10

# 2. H열에 총점(sum 이용), I열에 성적 정보 추가
#    총점 90이상 A, 80점 이상 B, 70점 이상 C, 나머지 D
ws['h1'].value = '합계'
ws['i1'].value = '성적 정보'
max_col += 2

# 3. 성적 정보 열에 A, B, C, D, F를 입력 
#    B:G열의 합계값을 구함 (openpyxl은 수식으로 부터 값을 구하지 못함)
for i, row in enumerate(ws.rows):
    if i == 0:  # 열 제목 행
        continue
    else:
        grade =''
        sum = 0
        for idx, cell in enumerate(row):
            if idx == 0:  # 제목 행을 의미
                continue
            elif idx == 1:   # 출석 정보 열을 의미 
                if cell.value < 5: 
                    grade = 'F'
            elif idx == 7:   # 합계 점수 열을 의미, H열에 '=sum( )' 수식을 입력
                formula_str = '=sum(b' + str(i + 1) + ':' + 'g' + str(i + 1) + ')'
                cell.value = formula_str
            elif idx == max_col - 1:  # 성적 정보 열을 의미 
                if not grade:
                    if sum >= 90:
                        grade = 'A'
                    elif sum >= 80:
                        grade = 'B'
                    elif sum >= 70:
                        grade = 'C'
                    else:
                        grade = 'D'
                cell.value = grade
            elif isinstance(cell.value, int):   # 합계를 계산함
                sum += cell.value

wb.save('quiz_scores.xlsx')
